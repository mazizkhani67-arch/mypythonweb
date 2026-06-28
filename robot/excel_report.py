import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
import os
from config import Config

class ExcelReporter:
    def __init__(self, db, reports_dir='reports'):
        self.db = db
        self.reports_dir = reports_dir
        
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir)
    
    def generate_weekly_report(self, group_id, week_start, week_end, group_name=""):
        """تولید گزارش هفتگی برای یک گروه خاص"""
        
        # دریافت داده‌های هفتگی
        data = self.db.get_weekly_data(
            group_id,
            week_start.isoformat(),
            week_end.isoformat()
        )
        
        if not data:
            return None
        
        # ایجاد کتاب کار جدید
        wb = openpyxl.Workbook()
        
        # ایجاد شیت‌های مختلف
        self._create_summary_sheet(wb, data, week_start, week_end, group_name)
        self._create_daily_sheet(wb, data)
        self._create_category_sheet(wb, data)
        self._create_user_sheet(wb, data)
        
        # حذف شیت پیش‌فرض
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # ذخیره فایل
        filename = f'weekly_report_{group_id}_{week_start.strftime("%Y%m%d")}_{week_end.strftime("%Y%m%d")}.xlsx'
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def _create_summary_sheet(self, wb, data, week_start, week_end, group_name):
        """ایجاد شیت خلاصه"""
        ws = wb.create_sheet('خلاصه', 0)
        
        # عنوان
        title = f'گزارش هزینه‌های هفتگی گروه {group_name}'
        if not group_name:
            title = f'گزارش هزینه‌های هفتگی'
        
        ws.merge_cells(f'A1:F1')
        cell = ws['A1']
        cell.value = f'{title}\n{week_start.strftime("%Y/%m/%d")} تا {week_end.strftime("%Y/%m/%d")}'
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        # هدرها
        headers = ['ردیف', 'تاریخ', 'کاربر', 'دسته‌بندی', 'مبلغ (تومان)', 'توضیحات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # داده‌ها
        for row_idx, record in enumerate(data, 4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3)
            ws.cell(row=row_idx, column=2, value=record[0])  # تاریخ
            ws.cell(row=row_idx, column=3, value=record[1])  # کاربر
            ws.cell(row=row_idx, column=4, value=record[2])  # دسته‌بندی
            ws.cell(row=row_idx, column=5, value=record[3])  # مبلغ
            ws.cell(row=row_idx, column=6, value=record[4])  # توضیحات
        
        # تنظیم عرض ستون‌ها
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # جمع کل
        total_row = len(data) + 4
        ws.cell(row=total_row, column=4, value='جمع کل:').font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=5)
        total_cell.value = f'=SUM(E4:E{total_row-1})'
        total_cell.font = Font(bold=True, color='FF0000')
        
        # تنظیمات زیبایی
        for row in range(3, total_row + 1):
            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    def _create_daily_sheet(self, wb, data):
        """ایجاد شیت گزارش روزانه"""
        ws = wb.create_sheet('گزارش روزانه')
        
        # گروه‌بندی بر اساس روز
        daily_data = {}
        for record in data:
            date_key = record[0]
            if date_key not in daily_data:
                daily_data[date_key] = []
            daily_data[date_key].append(record)
        
        row = 1
        for date_key, records in sorted(daily_data.items()):
            # عنوان روز
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = f'📅 {date_key}'
            cell.font = Font(size=14, bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            row += 1
            
            # هدر
            headers = ['کاربر', 'دسته‌بندی', 'مبلغ', 'توضیحات']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
            row += 1
            
            # داده‌های روز
            day_total = 0
            for record in records:
                ws.cell(row=row, column=1, value=record[1])  # کاربر
                ws.cell(row=row, column=2, value=record[2])  # دسته‌بندی
                ws.cell(row=row, column=3, value=record[3])  # مبلغ
                ws.cell(row=row, column=4, value=record[4])  # توضیحات
                day_total += record[3]
                row += 1
            
            # جمع روز
            ws.cell(row=row, column=2, value='جمع روز:').font = Font(bold=True)
            ws.cell(row=row, column=3, value=day_total).font = Font(bold=True, color='FF0000')
            row += 2
    
    def _create_category_sheet(self, wb, data):
        """ایجاد شیت دسته‌بندی"""
        ws = wb.create_sheet('دسته‌بندی')
        
        # گروه‌بندی بر اساس دسته‌بندی
        category_data = {}
        for record in data:
            category = record[2]
            if category not in category_data:
                category_data[category] = []
            category_data[category].append(record)
        
        row = 1
        ws.cell(row=row, column=1, value='دسته‌بندی').font = Font(bold=True, size=14)
        ws.cell(row=row, column=2, value='تعداد').font = Font(bold=True, size=14)
        ws.cell(row=row, column=3, value='مجموع مبلغ').font = Font(bold=True, size=14)
        row += 1
        
        for category, records in sorted(category_data.items()):
            total = sum(r[3] for r in records)
            count = len(records)
            
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=total)
            row += 1
    
    def _create_user_sheet(self, wb, data):
        """ایجاد شیت کاربران"""
        ws = wb.create_sheet('کاربران')
        
        # گروه‌بندی بر اساس کاربر
        user_data = {}
        for record in data:
            user = record[1]
            if user not in user_data:
                user_data[user] = []
            user_data[user].append(record)
        
        row = 1
        ws.cell(row=row, column=1, value='کاربر').font = Font(bold=True, size=14)
        ws.cell(row=row, column=2, value='تعداد').font = Font(bold=True, size=14)
        ws.cell(row=row, column=3, value='مجموع مبلغ').font = Font(bold=True, size=14)
        row += 1
        
        for user, records in sorted(user_data.items()):
            total = sum(r[3] for r in records)
            count = len(records)
            
            ws.cell(row=row, column=1, value=user)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=total)
            row += 1